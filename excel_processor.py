import os
import json
import re
from datetime import date, datetime

import pandas as pd
import matplotlib.pyplot as plt

from math import log10, floor

from openpyxl.utils import column_index_from_string, get_column_letter
# za proveru boje redova - zuti redovi su vazniji u fin listu
from openpyxl import load_workbook
from openpyxl.styles import Color
from unidecode import unidecode


from openai import OpenAI

def get_cell_value(df, cell_address):
    col_letter = ''.join(filter(str.isalpha, cell_address))
    row_number = int(''.join(filter(str.isdigit, cell_address))) - 1 # excel krece numeraciju od 1

    col_index = column_index_from_string(col_letter) - 1 # excel krece numeraciju od 1
    return df.iloc[row_number, col_index]


def find_header(file_path, sheet_name, column, range, keyword=""):
  try:
    df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=f"{column}:{column}", skiprows=range[0]-1, nrows=range[1]-range[0], engine='openpyxl', header=None)
  except Exception as e:
    #print(e)
    #print("KOLONA")
    return -1
  #print(df)
  if not df.empty:
    if keyword != "":
      for i, value in enumerate(df.iloc[:, 0]):
        if str(value).strip() == str(keyword).strip(): 
          return range[0] + i
    else:
        for i, value in enumerate(df.iloc[:, 0]):
          if str(value).strip().lower() != "nan":
            return range[0] + i
      
  return -1 

def find_yellow_rows(file_path):
  wb = load_workbook(file_path)
  ws = wb['Fin']
  zuta_boja = "FFFFFF00"

  zuti_redovi_excel = []

  for row in ws.iter_rows(min_row=2):  # preskačemo header red
      for cell in row:
          fill = cell.fill
          # Proveravamo da li boja postoji i da li je tip boje "rgb"
          if fill.start_color.type == "rgb" and fill.start_color.rgb == zuta_boja:
              zuti_redovi_excel.append(cell.row)
              break  # ne moramo proveravati ostale ćelije u tom redu

  #print("Žuti redovi su:", zuti_redovi_excel)
  return zuti_redovi_excel

def parse_date(d):
    if pd.isna(d):
        return None
    try:
        return pd.to_datetime(d).strftime('%Y-%m-%d')
    except:
        return str(d)


def read_excel_segment(file_path, sheet_name, usecols=None, skiprows=None, nrows=None, header=0, drop_all_nan=True):
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        usecols=usecols,
        skiprows=skiprows,
        nrows=nrows,
        engine='openpyxl',
        header=header
    )
    if drop_all_nan:
        df = df.dropna(how='all')
    return df.astype(object).where(pd.notnull(df), None)


def extract_osnovne_informacije(file_path, sheet_name):
  df_osnovne_informacije = read_excel_segment(file_path, sheet_name=sheet_name, usecols="A:A", skiprows=1, nrows=10, header=None)
  df_osnovne_informacije = df_osnovne_informacije.dropna()
  osnovne_informacije_json = {}
  if len(df_osnovne_informacije.iloc[0, 0]) > 3:
    osnovne_informacije_json['naziv_komitenta'] = df_osnovne_informacije.iloc[0, 0]
    osnovne_informacije_json['status_pravnog_lica'] = df_osnovne_informacije.iloc[1, 0].replace('\xa0', '')
    osnovne_informacije_json['delatnost'] = df_osnovne_informacije.iloc[3, 0].split(':')[1]
  else:
     osnovne_informacije_json['scoring'] = df_osnovne_informacije.iloc[0, 0]
     osnovne_informacije_json['iznosEUR'] = df_osnovne_informacije.iloc[1, 0]
     osnovne_informacije_json['naziv_komitenta'] = df_osnovne_informacije.iloc[2, 0]
     osnovne_informacije_json['status_pravnog_lica'] = df_osnovne_informacije.iloc[3, 0].replace('\xa0', '')
     osnovne_informacije_json['delatnost'] = df_osnovne_informacije.iloc[5, 0].split(':')[1]
  #  osnovne_informacije_json['maticni_broj'] = re.search(r"Matični broj:\s*(\d+)", df_osnovne_informacije.iloc[4, 0]).group(1)
  #  osnovne_informacije_json['pib'] = re.search(r"PIB:\s*(\d+)", df_osnovne_informacije.iloc[4, 0]).group(1)
  #  osnovne_informacije_json['pdv_status'] = re.search(r"PDV status:\s*([A-Z]+)", df_osnovne_informacije.iloc[4, 0]).group(1)
  #  ulica = re.search(r"Adresa:\s*([^\s].*?)\s{2,}", df_osnovne_informacije.iloc[6, 0]).group(1)
  #  mesto = re.search(r"Mesto:\s*(.+)", df_osnovne_informacije.iloc[6, 0]).group(1).replace('\xa0', '')
  #  osnovne_informacije_json['adresa'] = {
  #       "ulica": ulica,
  #       "mesto": mesto
  #  }
  #  osnovne_informacije_json['delatnost'] = re.search(r"Delatnost:\s*(.+?)\s{2,}", df_osnovne_informacije.iloc[5, 0]).group(1)
  #  osnovne_informacije_json['telefon'] = re.search(r"Telefon:\s*(.+)", df_osnovne_informacije.iloc[7, 0]).group(1).replace('\xa0', '')
  #  match = re.search(r"Blokade:\s*(.+)", df_osnovne_informacije.iloc[8, 0])
  #  osnovne_informacije_json['blokade'] = match.group(1).replace('\xa0', '') if match else None
  #  match = re.search(r"Izvršenja:\s*(.+)", df_osnovne_informacije.iloc[8, 0])
  #  osnovne_informacije_json['izvrsenja'] = match.group(1).replace('\xa0', '') if match else None

  return osnovne_informacije_json


def extract_poslovanje_po_godinama(file_path, sheet_name):
  h = find_header(file_path, sheet_name, "A", (11, 19), "Godina")
  df_poslovanje = read_excel_segment(file_path, sheet_name=sheet_name, skiprows=h-1, nrows=7, header=0, drop_all_nan=False)
  df_poslovanje = df_poslovanje.dropna(axis=1, how='all')
  #print(df_poslovanje)
  poslovanje_json = df_poslovanje.to_dict(orient='records')
  poslovanje_po_godinama = {}

  for red in poslovanje_json:
      metrika = red['Godina'].lower().replace(' ', '_')
      for godina in df_poslovanje.columns[1:]:
          if godina not in poslovanje_po_godinama:
              poslovanje_po_godinama[godina] = {}
          poslovanje_po_godinama[godina][metrika] = red.get(godina)

  return poslovanje_po_godinama, h


def extract_osnivanje_firme(file_path, sheet_name, poslovanje_header_row):
  df_osnivanje = read_excel_segment(file_path, sheet_name=sheet_name, usecols="A:B", skiprows=poslovanje_header_row+7+1, nrows=4, header=None, drop_all_nan=False)
  df_osnivanje.columns = ['Atribut', 'Vrednost']
  osnivanje_json = df_osnivanje.to_dict(orient='records')
  osnivanje_firme = {}
  for red in osnivanje_json:
      key = red['Atribut'].lower().replace(' ', '_')
      value = red['Vrednost']
      if isinstance(value, datetime):
          value = value.strftime('%Y-%m-%d')
      osnivanje_firme[key] = value

  return osnivanje_firme


def extract_blokade(file_path, sheet_name):
  h = find_header(file_path, sheet_name, "A", (13, 18), "Blokade računa")
  if h == -1:
    h = find_header(file_path, sheet_name, "A", (14, 18))
    blokada_od2010 = read_excel_segment(file_path, sheet_name=sheet_name, usecols="A:A", skiprows=h-1, nrows=1, header=None, drop_all_nan=False).iloc[0,0]
    return blokada_od2010

  df_raw = read_excel_segment(file_path, sheet_name=sheet_name, usecols="A:D", skiprows=h, header=0, drop_all_nan=False)
  stop_index = df_raw[df_raw.iloc[:, 0].str.contains("Računi", na=False)].index[0]
  df_blokada = df_raw.iloc[:stop_index].copy()

  df_blokada.reset_index(drop=True, inplace=True)
  df_blokada.dropna(how='all', inplace=True)
  df_blokada = df_blokada.astype(str)
  return df_blokada.to_dict(orient='records')


def extract_uvoz_izvoz(file_path, sheet_name):
  uvoz_json = None
  izvoz_json = None

  h = find_header(file_path, sheet_name, "A", (13, 18), "Godina")
  if h != -1:
    uvoz = read_excel_segment(file_path, sheet_name=sheet_name, usecols="A:F", skiprows=h-1, header=0)
    uvoz.columns = uvoz.columns.str.replace('\xa0', '', regex=False).str.strip()
    if ('Godina' in uvoz.columns) and any(uvoz['Godina'].apply(lambda x: isinstance(x, (int, float)))):
        uvoz_filtered = uvoz[uvoz['Godina'].apply(lambda x: isinstance(x, (int, float)))].copy()
        uvoz_filtered.reset_index(drop=True, inplace=True)
        uvoz_json = uvoz_filtered.to_dict(orient='records')

  h = find_header(file_path, sheet_name, "H", (13, 18), "Godina")
  if h != -1:
    izvoz = read_excel_segment(file_path, sheet_name=sheet_name, usecols="H:L", skiprows=h-1, header=0)
    izvoz.columns = izvoz.columns.str.replace('\xa0', '', regex=False).str.strip()
    if ('Godina.1' in izvoz.columns)and any(izvoz['Godina.1'].apply(lambda x: isinstance(x, (int, float)))):
        izvoz_filtered = izvoz[izvoz['Godina.1'].apply(lambda x: isinstance(x, (int, float)))].copy()
        izvoz_filtered.reset_index(drop=True, inplace=True)
        izvoz_json = izvoz_filtered.to_dict(orient='records')

  return uvoz_json, izvoz_json


def extract_apr_zabeleske(file_path, sheet_name):
  h = find_header(file_path, sheet_name, "A", (15, 20), "Tip")
  if h == -1:
    return []

  apr_zabeleske = read_excel_segment(file_path, sheet_name=sheet_name, usecols="A:C", skiprows=h-1, header=0, drop_all_nan=False)
  apr_zabeleske_clean = apr_zabeleske.dropna(how='all').copy()
  if 'Tip' not in apr_zabeleske_clean.columns:
    return []

  apr_zabeleske_clean = apr_zabeleske_clean[apr_zabeleske_clean['Tip'].notna()]
  apr_zabeleske_clean = apr_zabeleske_clean[apr_zabeleske_clean['Tip'] != '2.1.4.0']
  apr_zabeleske_clean.reset_index(drop=True, inplace=True)

  nezeljene_poruke = [
      "Nije pronadjena ni jedna ostala zabeležba.",
      "Nije pronadjena ni jedna obrisana APR zabeležba."
  ]
  apr_zabeleske_filtered = apr_zabeleske_clean[~apr_zabeleske_clean['Tip'].isin(nezeljene_poruke)]

  if not apr_zabeleske_filtered.empty and 'Datum' in apr_zabeleske_filtered.columns:
      apr_zabeleske_filtered['Datum'] = pd.to_datetime(
            apr_zabeleske_filtered['Datum'], dayfirst=True,
            format="%d.%m.%Y",
            errors='coerce'
        ).dt.strftime('%Y-%m-%d')
      return apr_zabeleske_filtered.to_dict(orient='records')

  return []


def extract_saradnja(file_path, sheet_name):
  seme = read_excel_segment(file_path, sheet_name=sheet_name, usecols="B:E", skiprows=3, nrows=6, header=0, drop_all_nan=False)
  pesticidi = read_excel_segment(file_path, sheet_name=sheet_name, usecols="B:E", skiprows=11, nrows=6, header=0, drop_all_nan=False)
  mehanizacija = read_excel_segment(file_path, sheet_name=sheet_name, usecols="B:E", skiprows=19, nrows=6, header=0, drop_all_nan=False)
  stocarstvo = read_excel_segment(file_path, sheet_name=sheet_name, usecols="B:E", skiprows=27, nrows=6, header=0, drop_all_nan=False)
  berza = read_excel_segment(file_path, sheet_name=sheet_name, usecols="B:E", skiprows=35, nrows=6, header=0, drop_all_nan=False)
  voce_povrce = read_excel_segment(file_path, sheet_name=sheet_name, usecols="B:E", skiprows=43, nrows=6, header=0, drop_all_nan=False)
  ratarstvo = read_excel_segment(file_path, sheet_name=sheet_name, usecols="B:E", skiprows=52, nrows=6, header=0, drop_all_nan=False)

  saradnja_sektori_rsd = {
    'seme': seme.to_dict(orient='records'),
    'pesticidi': pesticidi.to_dict(orient='records'),
    'mehanizacija': mehanizacija.to_dict(orient='records'),
    'stocarstvo': stocarstvo.to_dict(orient='records'),
    'berza': berza.to_dict(orient='records'),
    'voce_povrce': voce_povrce.to_dict(orient='records'),
    'ratarstvo': ratarstvo.to_dict(orient='records')
  }

  otvorene_stavke = read_excel_segment(file_path, sheet_name=sheet_name, usecols="G:M", skiprows=3, nrows=8, header=0, drop_all_nan=False)
  otvorene_stavke.columns = ["Sektor"] + list(otvorene_stavke.columns[1:])
  otvorene_stavke_json = otvorene_stavke.to_dict(orient='records')
  otvorene_stavke_filtered = [row for row in otvorene_stavke_json if str(row.get('Sektor', '')).lower() != 'grand total']

  return {
    'saradnja_sektori_rsd': saradnja_sektori_rsd,
    'otvorene_stavke': otvorene_stavke_filtered
  }
def extract_ebitda_from_sheet8(file_path, sheet_index=7, max_scan_rows=400, max_scan_cols=60):
    """
    Traži 'EBITDA' na 8. listu (sheet_index=7) bez oslanjanja na fiksan red.
    Vraća dict sa vrednostima iz ćelija desno od naziva (npr. za više godina).
    """
    df = read_excel_segment(
        file_path,
        sheet_name=sheet_index,
        header=None,
        drop_all_nan=False
    )

    if df.empty:
        return None

    # ograniči sken da bude brže (možeš povećati ako treba)
    df = df.iloc[:min(max_scan_rows, df.shape[0]), :min(max_scan_cols, df.shape[1])]

    # normalizuj sve u string za pretragu
    def norm(x):
        if x is None:
            return ""
        s = unidecode(str(x))
        s = re.sub(r"\s+", " ", s).strip().lower()
        return s

    # 1) nađi ćeliju gde piše EBITDA
    target_pos = None
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            if re.search(r"\bebitda\b", norm(df.iat[r, c]), flags=re.IGNORECASE):
                target_pos = (r, c)
                break
        if target_pos:
            break

    if not target_pos:
        return None

    r, c = target_pos

    # 2) pokušaj da uzmeš brojeve desno (npr. godine)
    # uzimamo do 10 kolona desno, filtriramo samo numeričke
    raw_right = []
    for cc in range(c + 1, min(df.shape[1], c + 11)):
        raw_right.append(df.iat[r, cc])

    # konverzija (podržava "1.234,56" format)
    values = []
    for v in raw_right:
        if v is None:
            continue
        if isinstance(v, (int, float)):
            values.append(float(v))
            continue
        s = str(v).strip()
        if s == "":
            continue
        s2 = s.replace(".", "").replace(",", ".")
        try:
            values.append(float(s2))
        except ValueError:
            # ignoriši ne-numeričke
            continue

    # 3) pokušaj da pronađeš godinu skeniranjem nagore po istoj koloni,
    # jer poslednji sheet nije strogo standardizovan
    def parse_numeric(raw):
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return float(raw)
        s = str(raw).strip()
        if not s:
            return None
        s2 = s.replace(".", "").replace(",", ".")
        try:
            return float(s2)
        except ValueError:
            return None

    def find_year_for_column(col_idx, start_row):
        for rr in range(start_row - 1, -1, -1):
            candidate = df.iat[rr, col_idx]
            parsed = parse_numeric(candidate)
            if parsed is not None and 1900 < int(parsed) < 2100:
                return int(parsed)
        return None

    years = []
    scanned_cols = list(range(c + 1, min(df.shape[1], c + 11)))
    for cc in scanned_cols:
        years.append(find_year_for_column(cc, r))

    exchange_rate = None
    for rr in range(min(5, df.shape[0])):
        for cc in range(max(0, c - 3), min(df.shape[1], c + 6)):
            label = norm(df.iat[rr, cc])
            if label in {"ex.rate", "ex rate", "kurs", "курс"}:
                for rate_col in range(cc + 1, min(df.shape[1], cc + 6)):
                    parsed = parse_numeric(df.iat[rr, rate_col])
                    if parsed is not None and parsed > 0:
                        exchange_rate = parsed
                        break
            if exchange_rate is not None:
                break
        if exchange_rate is not None:
            break

    # spakuj rezultat
    result = {
        "sheet_index": sheet_index,
        "found_at": {"row0": r, "col0": c},
        "values_right": values,
        "exchange_rate": exchange_rate,
    }

    # ako imamo godine (bar jednu), mapiraj year->value po poziciji
    if any(y is not None for y in years):
        year_value = {}
        # values je “zgusnut” (preskače None i tekst), pa je bolje mapirati po poziciji bez filtriranja:
        year_value_positional = {}
        latest_year = None
        latest_value = None
        for i, cc in enumerate(scanned_cols):
            y = years[i] if i < len(years) else None
            v = df.iat[r, cc]
            parsed = parse_numeric(v)
            if y is not None:
                year_value_positional[str(y)] = parsed
            if parsed is not None:
                latest_value = parsed
                if y is not None:
                    latest_year = str(y)
        result["by_year"] = year_value_positional
        result["latest_year"] = latest_year
        result["latest_value"] = latest_value
    else:
        # fallback: uzmi poslednju numeričku vrednost desno od EBITDA reda
        latest_value = None
        for cc in scanned_cols:
            parsed = parse_numeric(df.iat[r, cc])
            if parsed is not None:
                latest_value = parsed
        result["latest_year"] = None
        result["latest_value"] = latest_value

    return result

def extract_finansije(file_path):
  fin_blocks = {
    'bilans_stanja': (9, 128),
    'bilans_uspeha': (129, 193),
    'bilans_novcanih_troskova': (194, 252),
    'racio_analiza_likvidnosti': (254, 264),
    'racio_analiza_rentabilnost': (265, 271),
    'racio_analiza_aktivnosti': (272, 288)

  }

  zuti_redovi_excel = find_yellow_rows(file_path)
  df_fin = read_excel_segment(file_path, sheet_name='Fin', header=None, drop_all_nan=False)
  if df_fin.empty:
    error_message = f'Obavezni list "Fin" je prazan.'
    print(error_message)
    raise ValueError(error_message)

  n_years = min(3, df_fin.shape[1] - 1)
  fin_json = {}
  years_row_idx = 3
  years = df_fin.iloc[years_row_idx].dropna().tolist()
  zuti_redovi_df = [r - 1 for r in zuti_redovi_excel]
  df_fin = df_fin.iloc[:, :(len(years)+1)]

  for blok, (start_row, end_row) in fin_blocks.items():
    df_block = df_fin.iloc[start_row:end_row+1].copy()

    df_block['zuti_pokazatelj'] = df_block.index.isin(zuti_redovi_df)

    selected_years = years[-n_years:]

    n_cols = df_block.shape[1]
    df_final = df_block.iloc[:, [0] + list(range(n_cols - (n_years + 1), n_cols))]

    header = ['naziv'] + selected_years + ['zuti_pokazatelj']
    df_final.columns = header

    fin_json[blok] = df_final.to_dict(orient='records')

  return fin_json


def extract_ugovori(file_path, sheet_index):
  df_ugovori = read_excel_segment(file_path, sheet_name=sheet_index, usecols="W:AA", skiprows=14, nrows=10, header=0)
  df_ugovori['Zaključen'] = df_ugovori['Zaključen'].apply(parse_date)
  df_ugovori['Ispunjenje'] = df_ugovori['Ispunjenje'].apply(parse_date)
  return df_ugovori.to_dict(orient='records')


def extract_pokazatelji(file_path, sheet_index):
  df_pokazatelji = read_excel_segment(file_path, sheet_name=sheet_index, usecols="S:Y", skiprows=31, nrows=6, header=0)
  df_pokazatelji.columns = [unidecode(str(col)) for col in df_pokazatelji.columns]

  for col in df_pokazatelji.select_dtypes(include=['object']).columns:
      df_pokazatelji[col] = df_pokazatelji[col].apply(lambda x: unidecode(str(x)) if pd.notna(x) else x)

  return df_pokazatelji.to_dict(orient='records')
    

def to_JSON(file_path):
  final_json = {}

  # saradnja potencijalno prazna
  # 8. sheet je imenovan kao naziv firme
  sheet_names = ['OsnP', 'Blok', 'ImEx', 'Zabel', 'Fin', 'Saradnja']

  # osnovne informacije
  print("Obrada osnovnih informacija...")
  osnovne_informacije_json = extract_osnovne_informacije(file_path, sheet_names[0])
  final_json['osnovne_informacije'] = osnovne_informacije_json
  #print(f"Komitent {final_json['osnovne_informacije']['naziv_komitenta']}")
  

  # osnovno poslovanje po godinama
  print("Obrada poslovanja po godinama...")
  poslovanje_po_godinama, poslovanje_header = extract_poslovanje_po_godinama(file_path, sheet_names[0])
  final_json['poslovanje_po_godinama_osnovno'] = poslovanje_po_godinama


  # osnivanje
  print("Obrada detalja osnivanja...")
  final_json['osnivanje_firme'] = extract_osnivanje_firme(file_path, sheet_names[0], poslovanje_header)

  # blokada
  print("Obrada blokada...")
  final_json['blokade_od_2010'] = extract_blokade(file_path, sheet_names[1])
  

  # uvoz i izvoz – informativan podatak, može blago poboljšati procenu
  print("Obrada uvoza i izvoza...")
  try:
      uvoz_json, izvoz_json = extract_uvoz_izvoz(file_path, sheet_names[2])
      final_json['uvoz'] = uvoz_json
      final_json['izvoz'] = izvoz_json
  except Exception as e:
      print(f"Uvoz/izvoz nije dostupan: {e}")
      final_json['uvoz'] = None
      final_json['izvoz'] = None


  # APR zabeleske
  print("Obrada APR zabeleški...")
  final_json['apr_zabeleske'] = extract_apr_zabeleske(file_path, sheet_names[3])
  # Zadržan i stari ključ zbog postojećih pravila hard stop provere.
  final_json['zabelezbe'] = final_json['apr_zabeleske']

  #saradnja = extract_saradnja(file_path, sheet_names[5])
  #final_json['saradnja_sektori_rsd'] = saradnja['saradnja_sektori_rsd']
  #final_json['otvorene_stavke'] = saradnja['otvorene_stavke']

  # finansije
  print("Obrada finansije...")
  final_json['finansije'] = extract_finansije(file_path)

  # ugovori
  print("Obrada ugovori...")
  final_json['ugovori'] = extract_ugovori(file_path, 7)

  print("Obrada EBITDA...")
  final_json['ebitda'] = extract_ebitda_from_sheet8(file_path, sheet_index=7)

  # pokazatelji
  #print("Obrada pokazatelji...")
  #final_json['osnovni_pokazatelji'] = extract_pokazatelji(file_path, 7)

  return json.loads(json.dumps(final_json, ensure_ascii=False, default=str))


def propose_credit_limit(prihod_rsd: float,
                             a: float = 26.54,
                             b: float = -2.46,
                             lo: float = 0.1,
                             hi: float = 3.0) -> dict:
    """
    Računa procenat i predlog kreditnog limita na osnovu godišnjeg poslovnog prihoda.
    """
    if prihod_rsd <= 0:
        raise ValueError("Prihod mora biti pozitivan broj")

    pct = a + b * log10(prihod_rsd)
    pct = max(lo, min(hi, pct))  # clipping

    limit = prihod_rsd * pct / 100.0
    return {"proc_limita": pct, "predlog_limit": int(limit)}


def generate_AIcomment1(prompt, key):
  # Inicijalizuj klijenta
  client = OpenAI(api_key=key)  # Preporuka: koristi os.environ

  system_content = """
    You are an expert AI Credit Risk Analyst.
    Your task is to carefully analyze the provided client JSON data and generate a professional, ready-to-use "AI Comment" in business Serbian for a human credit risk analyst.
  """
  # Poziv modela (GPT-4.1)
  response = client.chat.completions.create(
      model="gpt-4.1",
      messages=[
          {"role": "system", "content": system_content},
          {"role": "user", "content": prompt}

      ],
      temperature=0.0
  )

  # Prikaz odgovora
  return response.choices[0].message.content

def generate_AIcomment(prompt, key):
  # Inicijalizuj klijenta
  client = OpenAI(api_key=key)  # Preporuka: koristi os.environ

  response = client.responses.create(
      model = "gpt-5-2025-08-07",
      input=prompt,
      reasoning={
          "effort": "high"
      },
      text={
          "verbosity": "high"
      },
      max_output_tokens= 15000
  )

  usage = response.usage

  usage_data = {
        "input_tokens": usage.input_tokens,
        "output_tokens": usage.output_tokens,
        "total_tokens": usage.total_tokens
    }
  
  return response.output_text, usage_data
